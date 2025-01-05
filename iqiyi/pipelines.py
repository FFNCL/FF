import os
import requests
import openpyxl
from scrapy.exceptions import DropItem
import re


class IqiyiPipeline:
    def __init__(self):
        print("开始初始化 IqiyiPipeline")

    def open_spider(self, spider):
        # 用于存储不同榜单对应的工作簿和工作表字典，键为榜单类型，值为对应的 (workbook, sheet) 元组
        self.rank_workbooks = {}
        for rank_type in ["0", "-6", "-5", "-4"]:
            folder_name = f"iqiyi_data_{rank_type}"
            if not os.path.exists(folder_name):
                os.makedirs(folder_name)
            print(f"获取到的文件夹名称为: {folder_name}")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # 设置表头，新增电影简介、说明信息、类型信息和热度信息对应的表头列
            sheet['A1'] = '电影名称'
            sheet['B1'] = '图片链接'
            sheet['C1'] = '电影简介'
            sheet['D1'] = '电影说明信息'
            sheet['E1'] = '电影类型'
            sheet['F1'] = '实时热度'
            self.rank_workbooks[rank_type] = (workbook, sheet)

    def process_item(self, item, spider):
        rank_type = item['rank_type']
        index = item['index']
        title = item['title']
        img_url = item.get('poster_url', '')
        introduction = item.get('introduction', '')  # 获取电影简介信息
        description = item.get('description', '')  # 获取电影说明信息
        movie_type = item.get('movie_type', '')  # 获取电影类型信息
        heat = item.get('heat', '')  # 获取电影实时热度信息

        # 检查标题、图片链接、简介、说明信息、类型信息和热度信息是否都存在，如果有缺失则抛出DropItem异常，不进行后续处理
        if not title or not img_url or not introduction or not description or not movie_type or not heat:
            raise DropItem(f"电影信息不完整，标题、图片链接、简介、说明信息、类型信息或热度信息缺失，标题: {title}，图片链接: {img_url}，简介: {introduction}，说明信息: {description}，类型: {movie_type}，热度: {heat}")

        workbook, sheet = self.rank_workbooks[rank_type]
        # 获取下一个空白行的行号，用于写入数据
        next_row = sheet.max_row + 1
        # 将电影名称、图片链接、电影简介、电影说明信息、电影类型和实时热度写入Excel表格对应的单元格
        sheet.cell(row=next_row, column=1, value=title)
        sheet.cell(row=next_row, column=2, value=img_url)
        sheet.cell(row=next_row, column=3, value=introduction)
        sheet.cell(row=next_row, column=4, value=description)
        sheet.cell(row=next_row, column=5, value=movie_type)
        sheet.cell(row=next_row, column=6, value=heat)

        # 尝试下载海报图片（如果图片链接有效）
        if img_url:
            try:
                img_response = requests.get(img_url)
                if img_response.status_code == 200:
                    # 处理电影名使其成为合法的文件名，替换掉不合法字符为下划线并去除首尾空白、替换空格为下划线
                    valid_title = re.sub(r'[^\w\s-]', '_', title).strip().replace(' ', '_')
                    img_file_path = os.path.join(f"iqiyi_data_{rank_type}", f"{index}-{valid_title}.jpg")
                    file_index = 1
                    while os.path.exists(img_file_path):
                        # 处理文件名重复时，添加序号的格式也保持数字.电影名的形式
                        img_file_path = os.path.join(f"iqiyi_data_{rank_type}", f"{index}-{file_index}_{valid_title}.jpg")
                        file_index += 1
                    with open(img_file_path, 'wb') as img_f:
                        img_f.write(img_response.content)
            except requests.RequestException as e:
                spider.logger.error(f"下载海报图片 {img_url} 时出错: {e}")

        return item

    def close_spider(self, spider):
        for rank_type, (workbook, _) in self.rank_workbooks.items():
            excel_file_path = os.path.join(f"iqiyi_data_{rank_type}", f"电影信息_{rank_type}.xlsx")
            workbook.save(excel_file_path)