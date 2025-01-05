import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 读取Excel文件
df = pd.read_excel('./豆瓣电影Top250.xlsx')

# 定义常见国家名称列表
common_countries = ['美国', '中国', '日本', '意大利', '法国', '英国', '印度', '韩国', '德国', '西班牙', '伊朗', '丹麦', '瑞典', '阿根廷', '泰国']

# 定义一个函数来提取国家名称
def extract_country(text):
    for country in common_countries:
        if country in text:
            return country
    return None

# 定义一个函数来提取电影类型，并去除剧情二字
def extract_genres(text):
    # 假设电影类型以逗号分隔
    genres = re.findall(r'\b(?:剧情|喜剧|动作|爱情|科幻|恐怖|悬疑|犯罪|冒险|奇幻|战争|历史|传记|动画|音乐|歌舞|家庭|西部|纪录片|短片|其他)\b', text)
    # 去除剧情类型
    genres = [genre for genre in genres if genre != '剧情']
    return ', '.join(genres) if genres else None

# 应用函数到 '相关信息' 列
df['国家'] = df['相关信息'].apply(extract_country)
df['类型'] = df['相关信息'].apply(extract_genres)

# 加载现有的Excel文件（如果存在）
try:
    wb = load_workbook('./豆瓣电影Top250_修改后.xlsx')
    # 检查是否存在名为'Processed_Data'的工作表，如果不存在则创建
    if 'Processed_Data' not in wb.sheetnames:
        ws = wb.create_sheet('Processed_Data')
    else:
        ws = wb['Processed_Data']
    # 清空工作表内容（除了表头，如果有）
    ws.delete_rows(2, ws.max_row)
except FileNotFoundError:
    # 如果文件不存在，则创建一个新的工作簿和工作表
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Processed_Data'

# 写入表头
headers = list(df.columns)
ws.append(headers)

# 逐行写入数据
for row in dataframe_to_rows(df, index=False, header=False):
    ws.append(row)

# 保存修改后的文件
wb.save('./豆瓣电影Top250_修改后.xlsx')