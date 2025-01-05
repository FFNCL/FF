import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def read_excel_file(file_path):
    try:
        df = pd.read_excel(file_path)
        if '电影类型' in df.columns:
            df['电影类型'] = df['电影类型'].astype(str)  # 确保电影类型列是字符串类型
        required_columns = {'电影类型', '实时热度'}
        missing_columns = required_columns - set(df.columns)
        if missing_columns:
            print(f"文件 {file_path} 缺少必要的列: {missing_columns}")
            return None
        return df
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到")
        return None
    except Exception as e:
        print(f"读取文件 {file_path} 时发生错误: {e}")
        return None


def split_and_clean_dataframe(df):
    if '电影类型' in df.columns:
        # 提取年份、电影类型和导演信息
        def extract_info(movie_type_str):
            parts = movie_type_str.split(' / ')
            if len(parts) < 3:
                return pd.Series([None] * 3, index=['年份', '电影类型_细分', '导演'])
            return pd.Series([parts[0], ' / '.join(parts[1:-1]), parts[-1]], index=['年份', '电影类型_细分', '导演'])

        # 应用函数提取电影类型相关信息
        new_columns = df['电影类型'].apply(extract_info)
        # 合并新提取的列与原数据框
        df = pd.concat([df, new_columns], axis=1)

        # 删除原始的电影类型列
        if '电影类型' in df.columns:
            df.drop('电影类型', axis=1, inplace=True)

        # 清理实时热度列中的非数字字符
        df['实时热度'] = df['实时热度'].apply(lambda x: re.sub(r'\D', '', str(x)))

    else:
        print("数据框中缺少电影类型列")
    return df


def process_files_in_directory():
    # 获取当前脚本所在目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    processed_data = {}

    for filename in os.listdir(current_dir):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(current_dir, filename)
            df = read_excel_file(file_path)
            if df is not None:
                try:
                    # 对数据框进行处理
                    processed_df = split_and_clean_dataframe(df)
                    processed_data[filename.rsplit('.', 1)[0]] = processed_df
                    print(f"成功处理文件: {file_path}")
                except Exception as e:
                    print(f"处理文件 {file_path} 时发生错误: {e}")

    if processed_data:
        # 创建一个新的Excel文件并加载
        output_path = os.path.join(current_dir, 'combined_processed_files.xlsx')
        try:
            wb = load_workbook(output_path)
            # 删除默认的第一个空白工作表
            if wb.active.title == 'Sheet':
                wb.remove(wb.active)
        except FileNotFoundError:
            from openpyxl import Workbook
            wb = Workbook()

        for sheet_name, df in processed_data.items():
            # 在Excel文件中创建新的工作表
            ws = wb.create_sheet(sheet_name)

            # 将数据框的列名写入工作表
            ws.append(df.columns.tolist())

            # 将数据框的每一行数据写入工作表
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

        # 保存Excel文件
        wb.save(output_path)
        print(f"所有文件处理完成，保存路径: {output_path}")
    else:
        print("没有成功处理任何文件")


if __name__ == "__main__":
    process_files_in_directory()